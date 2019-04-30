#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook 								#permite trabajar con archivos de excel
from openpyxl import load_workbook							#permite importar un archivo de excel
from openpyxl.styles import colors, borders, Font, Color, Fill,Border, Side, Alignment #Para el formato de los archivos
from openpyxl.styles.proxy import StyleProxy #Para estilos del archivo
from copy import copy #Permite copiar formato de variable a otro contexto
from PIL import Image #Permite insertar imagenes




class DocumentoExcel():
	"""docstring for DocumentoExcel:
	La presente clase sirve para hacer archivos con formato sencillo de Excel.
	"""

	__excelFile = "archivo01.xlsx"
	__fileAux = ""
	__filePath = ""

	def __init__(self, rutaArchivo):
		self.__filePath = rutaArchivo
		print("Ruta archivo->",rutaArchivo)
		print("Ruta self->",self.__filePath)

	def setGraphicStyle(self, celda, hoja, formato):
		if 'fuente' in formato:
			self.__excelFile[hoja][celda].font = Font(
				name=formato['fuente']['nombre'],
				bold=formato['bold'], 
				italic=formato['italic'], 
				color=formato['fuente']['color']
			)
			#self.__excelFile[hoja][celda].font = self.__excelFile[hoja][celda].font.copy()
		if 'fill' in formato:
			self.__excelFile[hoja][celda].fill = openpyxl.styles.PatternFill(
				start_color=formato['fill']['colorHex'], 
				end_color=formato['fill']['colorHex'], 
				fill_type = formato['fill']['tipo']
			)


			#print("aplicó fill")
		if 'size' in formato:
			self.__excelFile[hoja][celda].font = self.__excelFile[hoja][celda].font.copy(
				size=formato['size']
			)
			#print("aplicó size")
		if 'border' in formato:
			self.__excelFile[hoja][celda].border = openpyxl.styles.Border(
				left=Side(
					border_style=eval(formato['border']['tipo']),color=formato['border']['color']
					),
				right=Side(
					border_style=eval(formato['border']['tipo']),color=formato['border']['color']
					),
				top=Side(
					border_style=eval(formato['border']['tipo']),color=formato['border']['color']
					),
				bottom=Side(
					border_style=eval(formato['border']['tipo']),color=formato['border']['color'])
				)


		# Para mover la posición del texto en las celdas
		# default: horizontal='general', vertical='bottom', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
		if 'alignment' in formato:
			self.__excelFile[hoja][celda].alignment = openpyxl.styles.Alignment(
				horizontal=formato['alignment']['horiz'], # alineación sobre el eje x
				vertical=formato['alignment']['vert'], # alineación sobre el eje y
				text_rotation=formato['alignment']['rotation'], 
				wrap_text=formato['alignment']['wrap'],
				shrink_to_fit=formato['alignment']['shrink'],
				indent=formato['alignment']['indent']
				)

		return True

	'''
	hoja= String la hoja de excel a modificar
	sr= int renglon de inicio
	sc= int columna de inicio
	er = int renglon de fin
	ec=int columna de fin
	'''

	def mergeCeldas(self,hoja,sr,sc,er,ec):
		#sr,sc,er,ec deben ser int
		self.__excelFile[hoja].merge_cells(start_row=sr,start_column=sc,end_row=er,end_column=ec)
		return True


	def dimensionaCols(self,hoja,columna,medidas):
		# Letra de la columna va con Str
		self.__excelFile[hoja].column_dimensions[columna].width = medidas
		return True


	def dimensionaRows(self,hoja,fila,medidas):
		# Número del renglón va con Int
		self.__excelFile[hoja].row_dimensions[fila].height = medidas
		return True


	def escribeEnHoja(self, hoja, columna, row, texto, traceback=False, formato=None):
		"""
		Los parámetros recibidos: 
		hoja: str, 
		columna: str[A-Z], 
		row: int, 
		texto: str, 
		traceback: Boolean, 
		formato: dict
		"""
		if traceback == True:
			print('self.__excelFile[',hoja,"]","[",columna+str(row),"].value =",texto, "format:",formato)
		try:
			if formato is not None:
				celda = columna+str(row)
				self.setGraphicStyle(celda, hoja ,formato)
			self.__excelFile[ hoja ][columna+str(row)].value = texto
			return True
		except Exception as ex:
			print(ex)
			return False

	def abreArchivo(self, nameSheets, multiSheet=False, returnFile = False):
		"""
		nameSheets = [{'id':'reportEjec', 'name':'Reporte Ejecutivo'}] (Lista de diccionarios)
		"""
		self.__excelFile = Workbook()
		self.__excelFile.save(self.__filePath)

		self.__fileAux = load_workbook(self.__filePath)

		self.__excelFile = {}
		
		self.__excelFile[nameSheets[0]['id']] = self.__fileAux.active
		self.__excelFile[nameSheets[0]['id']].title = nameSheets[0]['name']

		if len(nameSheets) > 1:
			for a in range(1, len(nameSheets)):
				self.__excelFile[nameSheets[a]['id']] = self.__fileAux.create_sheet()
				self.__excelFile[nameSheets[a]['id']] = self.__fileAux['Sheet']
				self.__excelFile[nameSheets[a]['id']].title = nameSheets[a]['name']

		self.guardaArchivo()
		if returnFile == True:
			return self.__fileAux.active

	def addImageToCell(self, hoja, pathImage, cell):
		"""
		hoja: str
		pathImage: str
		cell: str
		"""
		img = openpyxl.drawing.image.Image(pathImage)
		self.__excelFile[hoja].add_image(img, cell)
		return True

	def guardaArchivo(self):
		#global roadmapFile
		#global filepathRoadMap
		self.__fileAux.save(self.__filePath)
		return True


	"""
	Ejemplo de formato: 
	formato = {
			'fuente': { 'nombre':"Calibri", 'color':'000000' },
			'bold'	: False,
			'italic': False,
			'fill'	: { 'colorHex':'008AD7', 'tipo':'solid'},
			'size'	: 11,
			'border': {'tipo':borders.BORDER_THIN, 'color':'000000'}
		}
	"""

